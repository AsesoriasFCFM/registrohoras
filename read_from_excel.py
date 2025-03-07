import pandas as pd
from models import session, Asesor, Entrada
from openpyxl import load_workbook
from datetime import datetime


def try_parse_date(date):
    """Try to parse a date string. Currently supports DD/MM/YYYY, DD/MM/YY, DD-MM-YYYY, DD-MM-YY and DD-MM."""
    for fmt in ["%d/%m/%Y", "%d/%m/%y", "%d/%m", "%d-%m-%Y", "%d-%m-%y", "%d-%m"]:
        try:
            return datetime.strptime(date, fmt)
        except ValueError:
            pass
    return None


def try_parse_time(time):
    """Try to parse a time string. Currently supports HH:MM and HH:MM:SS."""
    for fmt in ["%H:%M", "%H:%M:%S"]:
        try:
            return datetime.strptime(time, fmt)
        except ValueError:
            pass
    return None


def populate_db_from_excel():
    """Populate the database from the Excel file."""
    wb = load_workbook("Asesores.xlsx")
    ws_asesores = wb["Asesores"]

    for row in ws_asesores.iter_rows(min_row=2, values_only=True):
        asesor = Asesor(
            nombre=row[0],
            matricula=row[1],
            carrera=row[2],
            programa=row[3],
        )
        session.add(asesor)
    session.commit()

    for ws in wb.sheetnames:
        if ws == "Asesores":
            continue
        for row in wb[ws].iter_rows(min_row=2, values_only=True):
            asesor = session.query(Asesor).filter_by(matricula=row[1]).first()
            if not asesor:
                print(f"Matrícula '{row[1]}' no encontrada")
                continue
            entrada = Entrada(
                hora_entrada=try_parse_time(row[2]),
                hora_salida=try_parse_time(row[3]),
                fecha=try_parse_date(ws),
                horas_recuperadas=row[4],
                fecha_falta=try_parse_date(row[5]),
                asesor=asesor,
            )
            session.add(entrada)
        session.commit()
