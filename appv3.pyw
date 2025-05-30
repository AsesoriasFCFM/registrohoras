import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
import logging
import sys
from tkcalendar import DateEntry
import sqlite3
import calendar
import shutil # <--- AÑADIDO para backups
import glob   # <--- AÑADIDO para listar backups

NOMBRE_ARCHIVO_EXCEL = "Reporte_Asistencias.xlsx"
NOMBRE_BD = "datos_asesores.db"
DIRECTORIO_BACKUPS_BD = "backups_db" # <--- NUEVA CONSTANTE para backups

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    filename="app_asesores.log",
)
logger = logging.getLogger(__name__)


def capturar_excepcion(exc_type, exc_value, exc_traceback):
    logger.error(
        "Excepción no capturada", exc_info=(exc_type, exc_value, exc_traceback)
    )
    messagebox.showerror(
        "Error Inesperado",
        "Ha ocurrido un error inesperado.\nRevise app_asesores.log para más detalles.",
    )


sys.excepthook = capturar_excepcion


# --- Funciones de Validación para Entradas (validate='key') ---
def validar_solo_numeros_longitud(valor_nuevo, longitud_max):
    if valor_nuevo == "":
        return True
    if not valor_nuevo.isdigit():
        return False
    if len(valor_nuevo) > longitud_max:
        return False
    return True


def validar_horas_recuperar(valor_nuevo):
    if valor_nuevo == "":
        return True
    if not all(c.isdigit() or c == "." for c in valor_nuevo):
        return False
    if valor_nuevo.count(".") > 1:
        return False
    if "." in valor_nuevo and len(valor_nuevo.split(".")[1]) > 1:
        return False
    return True


def validar_mes(valor_nuevo):
    if valor_nuevo == "":
        return True
    if not valor_nuevo.isdigit():
        return False
    if len(valor_nuevo) > 2:
        return False
    return True


def validar_anio(valor_nuevo):
    if valor_nuevo == "":
        return True
    if not valor_nuevo.isdigit():
        return False
    if len(valor_nuevo) > 4:
        return False
    return True

# --- Funciones de Backup y Restauración ---
def crear_backup_bd_diario():
    """
    Crea un backup diario de la base de datos si no existe uno para el día actual,
    o si existe, lo sobrescribe (manteniendo solo el último del día).
    Los backups se guardan en el directorio especificado por DIRECTORIO_BACKUPS_BD.
    """
    if not os.path.exists(DIRECTORIO_BACKUPS_BD):
        try:
            os.makedirs(DIRECTORIO_BACKUPS_BD)
            logger.info(f"Directorio de backups creado: {DIRECTORIO_BACKUPS_BD}")
        except OSError as e:
            logger.error(f"No se pudo crear el directorio de backups {DIRECTORIO_BACKUPS_BD}: {e}")
            messagebox.showerror("Error de Backup", f"No se pudo crear el directorio de backups: {e}", parent=ventana if 'ventana' in globals() else None)
            return False

    fecha_hoy_str = datetime.now().strftime("%Y-%m-%d")
    nombre_archivo_backup = f"backup_bd_{fecha_hoy_str}.db"
    ruta_backup_destino = os.path.join(DIRECTORIO_BACKUPS_BD, nombre_archivo_backup)

    try:
        if os.path.exists(NOMBRE_BD):
            shutil.copy2(NOMBRE_BD, ruta_backup_destino)
            logger.info(f"Backup de la BD creado/actualizado: {ruta_backup_destino}")
            return True
        else:
            logger.warning(f"No se encontró el archivo de base de datos '{NOMBRE_BD}' para hacer backup.")
            return False
    except Exception as e:
        logger.error(f"Error al crear/actualizar el backup de la BD '{ruta_backup_destino}': {e}")
        messagebox.showerror("Error de Backup", f"No se pudo crear el backup de la base de datos: {e}", parent=ventana if 'ventana' in globals() else None)
        return False

def dialogo_restaurar_bd_desde_backup():
    if not os.path.exists(DIRECTORIO_BACKUPS_BD) or not os.listdir(DIRECTORIO_BACKUPS_BD):
        messagebox.showinfo("Restaurar BD", "No se encontraron backups disponibles.", parent=ventana)
        logger.info("Intento de restauración: No hay backups disponibles.")
        return

    patron_backup = os.path.join(DIRECTORIO_BACKUPS_BD, "backup_bd_*.db")
    lista_backups_full_path = sorted(glob.glob(patron_backup), reverse=True)

    if not lista_backups_full_path:
        messagebox.showinfo("Restaurar BD", "No se encontraron archivos de backup válidos (formato: backup_bd_YYYY-MM-DD.db).", parent=ventana)
        logger.info("Intento de restauración: No hay archivos de backup válidos.")
        return

    dialogo_seleccion = tk.Toplevel(ventana)
    dialogo_seleccion.title("Seleccionar Backup para Restaurar")
    dialogo_seleccion.geometry("450x350") # Un poco más de alto para el mensaje
    dialogo_seleccion.transient(ventana)
    dialogo_seleccion.grab_set()

    tk.Label(dialogo_seleccion, text="Seleccione el archivo de backup a restaurar:", pady=10).pack()
    
    tk.Label(dialogo_seleccion, text="Los backups se muestran del más reciente al más antiguo.", font=("Segoe UI", 8), fg="grey").pack()


    listbox_backups = tk.Listbox(dialogo_seleccion, width=60, height=10)
    for backup_path in lista_backups_full_path:
        listbox_backups.insert(tk.END, os.path.basename(backup_path))
    listbox_backups.pack(pady=5)
    if lista_backups_full_path:
        listbox_backups.select_set(0)

    seleccion_ruta_backup = tk.StringVar()

    def confirmar_seleccion():
        seleccion_indices = listbox_backups.curselection()
        if not seleccion_indices:
            messagebox.showwarning("Selección Requerida", "Por favor, seleccione un archivo de backup.", parent=dialogo_seleccion)
            return
        
        nombre_backup_seleccionado = listbox_backups.get(seleccion_indices[0])
        # Reconstruir la ruta completa del backup seleccionado
        ruta_completa_seleccionada = os.path.join(DIRECTORIO_BACKUPS_BD, nombre_backup_seleccionado)
        seleccion_ruta_backup.set(ruta_completa_seleccionada)
        dialogo_seleccion.destroy()

    btn_frame = tk.Frame(dialogo_seleccion)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="Restaurar Seleccionado", command=confirmar_seleccion).pack(side=tk.LEFT, padx=5)
    tk.Button(btn_frame, text="Cancelar", command=dialogo_seleccion.destroy).pack(side=tk.LEFT, padx=5)
    
    dialogo_seleccion.wait_window()

    ruta_backup_a_restaurar = seleccion_ruta_backup.get()

    if not ruta_backup_a_restaurar:
        logger.info("Restauración de BD cancelada por el usuario.")
        return

    advertencia = (
        "¡¡¡ADVERTENCIA EXTREMA!!!\n\n"
        "Está a punto de reemplazar la base de datos actual con el contenido del backup:\n"
        f"'{os.path.basename(ruta_backup_a_restaurar)}'.\n\n"
        "TODOS LOS DATOS ACTUALES NO GUARDADOS EN ESTE BACKUP SE PERDERÁN PERMANENTEMENTE.\n\n"
        "Se creará un backup de la base de datos actual ('..._antes_de_restaurar_...') antes de la restauración.\n\n"
        "¿Está ABSOLUTAMENTE SEGURO de que desea continuar?"
    )
    if not messagebox.askyesno("Confirmar Restauración de Base de Datos", advertencia, icon='warning', default=messagebox.NO, parent=ventana):
        logger.info(f"Restauración de BD desde '{ruta_backup_a_restaurar}' cancelada por el usuario tras advertencia.")
        return

    fecha_hora_actual_str = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    nombre_bd_actual_backup = f"{os.path.splitext(NOMBRE_BD)[0]}_antes_de_restaurar_{fecha_hora_actual_str}.db"
    ruta_bd_actual_backup = os.path.join(DIRECTORIO_BACKUPS_BD, nombre_bd_actual_backup)

    try:
        if os.path.exists(NOMBRE_BD):
            shutil.copy2(NOMBRE_BD, ruta_bd_actual_backup)
            logger.info(f"Backup de emergencia creado ANTES de la restauración: {ruta_bd_actual_backup}")
        else:
            logger.warning(f"La base de datos actual '{NOMBRE_BD}' no existe. No se creó backup de emergencia.")
    except Exception as e:
        logger.error(f"Error CRÍTICO al crear backup de emergencia ANTES de restaurar: {e}. Restauración ABORTADA.")
        messagebox.showerror("Error Crítico", f"No se pudo crear el backup de emergencia. Restauración abortada: {e}", parent=ventana)
        return

    try:
        shutil.copy2(ruta_backup_a_restaurar, NOMBRE_BD)
        logger.critical(f"RESTAURACIÓN DE BASE DE DATOS COMPLETADA. BD actual reemplazada con: '{ruta_backup_a_restaurar}'")
        messagebox.showinfo("Restauración Exitosa",
                            "La base de datos ha sido restaurada exitosamente.\n\n"
                            "La aplicación se reiniciará ahora para aplicar los cambios.", parent=ventana)
        # Forzar reinicio de la aplicación
        ventana.quit() # Cierra el bucle principal de Tkinter
        python = sys.executable
        os.execl(python, python, *sys.argv) # Vuelve a ejecutar el script actual

    except Exception as e:
        logger.error(f"Error CRÍTICO durante la restauración de la BD desde '{ruta_backup_a_restaurar}': {e}")
        messagebox.showerror("Error de Restauración", f"Ocurrió un error al restaurar la base de datos: {e}\n"
                                                    "La base de datos actual podría estar en un estado inconsistente. "
                                                    f"Revise el backup de emergencia: {ruta_bd_actual_backup}", parent=ventana)

# --- Funciones de Base de Datos SQLite ---
def inicializar_bd():
    conn = sqlite3.connect(NOMBRE_BD)
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS asesores (
            matricula TEXT PRIMARY KEY,
            nombre TEXT NOT NULL,
            carrera TEXT NOT NULL,
            programa TEXT NOT NULL,
            activo INTEGER DEFAULT 1 NOT NULL 
        )
    """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS registros_asistencia (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricula TEXT NOT NULL,
            hora_entrada TEXT,
            hora_salida TEXT,
            horas_recuperadas TEXT,
            fecha_falta_recuperada TEXT,
            fecha_registro TEXT NOT NULL,
            nota TEXT,
            FOREIGN KEY (matricula) REFERENCES asesores (matricula) ON UPDATE CASCADE ON DELETE RESTRICT
        )
    """
    )

    cursor.execute("PRAGMA table_info(asesores)")
    columnas_asesores = [info[1] for info in cursor.fetchall()]
    if "activo" not in columnas_asesores:
        cursor.execute(
            "ALTER TABLE asesores ADD COLUMN activo INTEGER DEFAULT 1 NOT NULL"
        )
        logger.info(
            "Columna 'activo' añadida a la tabla 'asesores' con valor por defecto 1."
        )

    cursor.execute("PRAGMA table_info(registros_asistencia)")
    columnas_registros = [info[1] for info in cursor.fetchall()]
    if "nota" not in columnas_registros:
        cursor.execute("ALTER TABLE registros_asistencia ADD COLUMN nota TEXT")
        logger.info("Columna 'nota' añadida a la tabla 'registros_asistencia'.")

    conn.commit()
    crear_backup_bd_diario() # <--- AÑADIDO backup después de inicializar/modificar estructura
    conn.close()
    logger.info("Base de datos inicializada/verificada.")


def obtener_conexion_bd():
    conn = sqlite3.connect(NOMBRE_BD)
    conn.row_factory = sqlite3.Row
    return conn


# --- Función para Regenerar el Excel desde la BD ---
def regenerar_excel_desde_bd(mostrar_mensaje_exito=False):
    logger.info(f"Regenerando reporte Excel: {NOMBRE_ARCHIVO_EXCEL}")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    conn = obtener_conexion_bd()
    cursor = conn.cursor()

    font_cabecera = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    fill_cabecera = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    alignment_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws_asesores = wb.create_sheet(title="Asesores Activos")
    cabeceras_asesores = ["Nombre", "Matrícula", "Carrera", "Programa"]
    ws_asesores.append(cabeceras_asesores)
    for col_idx, header_title in enumerate(cabeceras_asesores, 1):
        cell = ws_asesores.cell(row=1, column=col_idx)
        cell.font = font_cabecera
        cell.fill = fill_cabecera
        cell.alignment = alignment_centro
        cell.border = thin_border

    cursor.execute(
        "SELECT nombre, matricula, carrera, programa FROM asesores WHERE activo = 1 ORDER BY programa, carrera, nombre"
    )
    for idx, row_data in enumerate(cursor.fetchall(), 2):
        ws_asesores.append(
            [
                row_data["nombre"],
                row_data["matricula"],
                row_data["carrera"],
                row_data["programa"],
            ]
        )
        for col_idx in range(1, len(cabeceras_asesores) + 1):
            ws_asesores.cell(row=idx, column=col_idx).border = thin_border
            ws_asesores.cell(row=idx, column=col_idx).alignment = Alignment(
                vertical="center", wrap_text=True
            )
    logger.info("Hoja 'Asesores Activos' generada.")

    cursor.execute(
        "SELECT DISTINCT fecha_registro FROM registros_asistencia ORDER BY fecha_registro DESC"
    )
    fechas_distintas = [row["fecha_registro"] for row in cursor.fetchall()]
    cabeceras_diarias = [
        "Nombre Asesor",
        "Matrícula",
        "Hora de Entrada",
        "Hora de Salida",
        "Horas Trabajadas",
        "Horas Recuperadas",
        "Fecha Falta (Recup.)",
        "Nota",
        "Carrera",
        "Programa",
    ]
    for fecha_registro_str in fechas_distintas:
        try:
            dt_obj = datetime.strptime(fecha_registro_str, "%Y-%m-%d")
            titulo_hoja = dt_obj.strftime("%d-%m-%Y")
        except ValueError:
            titulo_hoja = fecha_registro_str

        ws_dia = wb.create_sheet(title=titulo_hoja)
        ws_dia.append(cabeceras_diarias)
        for col_idx, header_title in enumerate(cabeceras_diarias, 1):
            cell = ws_dia.cell(row=1, column=col_idx)
            cell.font = font_cabecera
            cell.fill = fill_cabecera
            cell.alignment = alignment_centro
            cell.border = thin_border

        cursor.execute(
            """
            SELECT ra.*, a.nombre, a.carrera, a.programa
            FROM registros_asistencia ra
            JOIN asesores a ON ra.matricula = a.matricula 
            WHERE ra.fecha_registro = ? 
            ORDER BY ra.hora_entrada, ra.matricula 
        """,
            (fecha_registro_str,),
        )

        for idx, row_data in enumerate(cursor.fetchall(), 2):
            horas_trabajadas_str = ""
            if row_data["hora_entrada"] and row_data["hora_salida"]:
                try:
                    dt_entrada = datetime.strptime(row_data["hora_entrada"], "%H:%M:%S")
                    dt_salida = datetime.strptime(row_data["hora_salida"], "%H:%M:%S")
                    diff = (dt_salida - dt_entrada).total_seconds()
                    if diff < 0:
                        diff += 86400
                    h = int(diff // 3600)
                    m = int((diff % 3600) // 60)
                    s = int(diff % 60)
                    horas_trabajadas_str = f"{h:02d}:{m:02d}:{s:02d}"
                except ValueError:
                    horas_trabajadas_str = "Error Calc."

            ws_dia.append(
                [
                    row_data["nombre"],
                    row_data["matricula"],
                    row_data["hora_entrada"],
                    row_data["hora_salida"],
                    horas_trabajadas_str,
                    row_data["horas_recuperadas"],
                    row_data["fecha_falta_recuperada"],
                    row_data["nota"],
                    row_data["carrera"],
                    row_data["programa"],
                ]
            )
            for col_idx in range(1, len(cabeceras_diarias) + 1):
                ws_dia.cell(row=idx, column=col_idx).border = thin_border
                ws_dia.cell(row=idx, column=col_idx).alignment = Alignment(
                    vertical="center", wrap_text=True
                )
        logger.info(f"Hoja de asistencia generada para fecha: {titulo_hoja}")
    conn.close()

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        for col in ws.columns:
            max_longitud = 0
            columna_letra = col[0].column_letter
            for celda in col:
                try:
                    if celda.value:
                        longitud_celda = len(str(celda.value))
                        if celda.row == 1 and celda.alignment.wrap_text:
                            longitud_celda = (
                                max(len(s) for s in str(celda.value).split())
                                if str(celda.value)
                                else 0
                            )
                        max_longitud = max(max_longitud, longitud_celda)
                except:
                    pass
            ancho_ajustado = (max_longitud + 3) if max_longitud > 0 else 10
            ws.column_dimensions[columna_letra].width = ancho_ajustado

    try:
        wb.save(NOMBRE_ARCHIVO_EXCEL)
        logger.info(f"Reporte Excel guardado exitosamente: {NOMBRE_ARCHIVO_EXCEL}.")
        if mostrar_mensaje_exito:
            messagebox.showinfo(
                "Reporte Actualizado",
                f"El archivo Excel '{NOMBRE_ARCHIVO_EXCEL}' ha sido actualizado.",
                parent=ventana if 'ventana' in globals() else None
            )
    except PermissionError:
        mensaje_error = (
            f"Permiso denegado al guardar '{NOMBRE_ARCHIVO_EXCEL}'.\n"
            f"Asegúrate de que no esté abierto en otro programa.\n\n"
            f"Puedes intentar actualizar el reporte manualmente usando el botón correspondiente "
            f"una vez que el archivo esté cerrado."
        )
        messagebox.showerror("Error al Guardar Excel", mensaje_error, parent=ventana if 'ventana' in globals() else None)
        logger.error(f"PermissionError al guardar {NOMBRE_ARCHIVO_EXCEL}.")
    except Exception as e:
        messagebox.showerror(
            "Error al Guardar Excel",
            f"No se pudo guardar el archivo Excel '{NOMBRE_ARCHIVO_EXCEL}': {e}",
            parent=ventana if 'ventana' in globals() else None
        )
        logger.error(
            f"Fallo al guardar Excel {NOMBRE_ARCHIVO_EXCEL}: {e}", exc_info=True
        )


# --- Funciones de la GUI ---
def limpiar_campos():
    entrada_matricula.delete(0, tk.END)
    entrada_horas_rec.delete(0, tk.END)
    entrada_fecha_falta_rec.set_date(datetime.today() - timedelta(days=1))
    entrada_nota.delete(0, tk.END)
    entrada_matricula.focus_set()


def registrar_entrada_accion(evento=None):
    matricula = entrada_matricula.get().strip()
    if not (matricula.isdigit() and len(matricula) == 7):
        messagebox.showerror("Error de Entrada", "La matrícula debe ser de 7 números.", parent=ventana)
        return

    conn = obtener_conexion_bd()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT nombre, carrera, programa FROM asesores WHERE matricula = ? AND activo = 1",
        (matricula,),
    )
    asesor = cursor.fetchone()
    if not asesor:
        messagebox.showerror(
            "Error de Matrícula",
            f"Matrícula '{matricula}' no encontrada o asesor inactivo.",
            parent=ventana
        )
        conn.close()
        return

    fecha_hoy_str = datetime.today().strftime("%Y-%m-%d")
    cursor.execute(
        "SELECT id FROM registros_asistencia WHERE matricula = ? AND fecha_registro = ? AND hora_salida IS NULL",
        (matricula, fecha_hoy_str),
    )
    if cursor.fetchone():
        messagebox.showwarning(
            "Registro Existente", "Ya existe una entrada abierta para este asesor hoy.", parent=ventana
        )
        conn.close()
        return

    hora_entrada_str = datetime.now().strftime("%H:%M:%S")
    nota_val = entrada_nota.get().strip()
    if not nota_val:
        nota_val = None

    horas_rec_val_str = entrada_horas_rec.get().strip()
    fecha_falta_val = None
    if horas_rec_val_str:
        try:
            horas_rec_float = float(horas_rec_val_str.replace(",", "."))
            if not (0 < horas_rec_float <= 8):
                messagebox.showerror(
                    "Entrada Inválida",
                    "Las horas a recuperar deben estar entre 0.1 y 8.",
                    parent=ventana
                )
                conn.close()
                return
            horas_rec_val_str = f"{horas_rec_float:.1f}"
            fecha_falta_val = entrada_fecha_falta_rec.get_date().strftime("%d/%m/%Y")
        except ValueError:
            messagebox.showerror(
                "Entrada Inválida", "Formato de horas a recuperar inválido.", parent=ventana
            )
            conn.close()
            return
    else:
        horas_rec_val_str = None

    try:
        cursor.execute(
            """
            INSERT INTO registros_asistencia 
            (matricula, hora_entrada, fecha_registro, horas_recuperadas, fecha_falta_recuperada, nota)
            VALUES (?, ?, ?, ?, ?, ?)
        """,
            (
                matricula,
                hora_entrada_str,
                fecha_hoy_str,
                horas_rec_val_str,
                fecha_falta_val,
                nota_val,
            ),
        )
        conn.commit()
        crear_backup_bd_diario() # <--- AÑADIDO backup después de registrar entrada
        mensaje_exito = f"Entrada registrada para {asesor['nombre']} ({asesor['carrera']} - {asesor['programa']}) a las {hora_entrada_str}"
        if nota_val:
            mensaje_exito += f"\nNota: {nota_val}"
        if horas_rec_val_str:
            mensaje_exito += (
                f"\nCon {horas_rec_val_str}h de recuperación para {fecha_falta_val}."
            )
        messagebox.showinfo("Registro Exitoso", mensaje_exito, parent=ventana)
        logger.info(
            f"Entrada: {matricula} @{hora_entrada_str}. Nota: '{nota_val or ''}'. Rec: {horas_rec_val_str or 'N/A'}"
        )
        limpiar_campos()
        regenerar_excel_desde_bd()
    except sqlite3.Error as e:
        messagebox.showerror(
            "Error de Base de Datos", f"No se pudo registrar la entrada: {e}", parent=ventana
        )
        logger.error(f"Error BD entrada {matricula}: {e}", exc_info=True)
    finally:
        conn.close()


def registrar_salida_accion(evento=None):
    if evento and evento.keysym not in ("Shift_L", "Shift_R"):
        return

    matricula = entrada_matricula.get().strip()
    if not (matricula.isdigit() and len(matricula) == 7):
        messagebox.showerror(
            "Error de Entrada",
            "La matrícula debe ser de 7 números para registrar salida.",
            parent=ventana
        )
        return

    conn = obtener_conexion_bd()
    cursor = conn.cursor()
    fecha_hoy_str = datetime.today().strftime("%Y-%m-%d")
    cursor.execute(
        """
        SELECT id, hora_entrada, a.nombre, a.carrera, a.programa
        FROM registros_asistencia ra JOIN asesores a ON ra.matricula = a.matricula
        WHERE ra.matricula = ? AND a.activo = 1 AND ra.fecha_registro = ? AND ra.hora_salida IS NULL
        ORDER BY ra.id DESC LIMIT 1
    """,
        (matricula, fecha_hoy_str),
    )
    registro_abierto = cursor.fetchone()
    if not registro_abierto:
        messagebox.showerror(
            "Error de Registro",
            "No se encontró una entrada pendiente para este asesor activo hoy.",
            parent=ventana
        )
        conn.close()
        return

    hora_salida_str = datetime.now().strftime("%H:%M:%S")
    nota_val = entrada_nota.get().strip()
    if not nota_val:
        nota_val = None

    horas_rec_val_str = entrada_horas_rec.get().strip()
    fecha_falta_val = None
    if horas_rec_val_str:
        try:
            horas_rec_float = float(horas_rec_val_str.replace(",", "."))
            if not (0 < horas_rec_float <= 8):
                messagebox.showerror(
                    "Entrada Inválida",
                    "Las horas a recuperar deben estar entre 0.1 y 8.",
                    parent=ventana
                )
                conn.close()
                return
            horas_rec_val_str = f"{horas_rec_float:.1f}"
            fecha_falta_val = entrada_fecha_falta_rec.get_date().strftime("%d/%m/%Y")
        except ValueError:
            messagebox.showerror(
                "Entrada Inválida", "Formato de horas a recuperar inválido.", parent=ventana
            )
            conn.close()
            return
    else:
        horas_rec_val_str = None

    try:
        update_fields = ["hora_salida = ?"]
        params = [hora_salida_str]
        if horas_rec_val_str is not None:
            update_fields.append("horas_recuperadas = ?")
            update_fields.append("fecha_falta_recuperada = ?")
            params.extend([horas_rec_val_str, fecha_falta_val])
        if nota_val is not None:
            update_fields.append("nota = ?")
            params.append(nota_val)

        query_str = (
            f"UPDATE registros_asistencia SET {', '.join(update_fields)} WHERE id = ?"
        )
        params.append(registro_abierto["id"])

        cursor.execute(query_str, tuple(params))
        conn.commit()
        crear_backup_bd_diario() # <--- AÑADIDO backup después de registrar salida

        dt_entrada = datetime.strptime(registro_abierto["hora_entrada"], "%H:%M:%S")
        dt_salida = datetime.strptime(hora_salida_str, "%H:%M:%S")
        diff = (dt_salida - dt_entrada).total_seconds()
        if diff < 0:
            diff += 86400
        h = int(diff // 3600)
        m = int((diff % 3600) // 60)
        s = int(diff % 60)

        msg = f"Salida registrada para {registro_abierto['nombre']}."
        msg += f"\nTiempo trabajado: {h:02d}:{m:02d}:{s:02d}."
        if nota_val:
            msg += f"\nNota: {nota_val}"
        if horas_rec_val_str:
            msg += f"\nHoras recuperación ({horas_rec_val_str}h) para {fecha_falta_val} también registradas/actualizadas."

        messagebox.showinfo("Registro Exitoso", msg, parent=ventana)
        logger.info(
            f"Salida: {matricula} @{hora_salida_str}. Nota: '{nota_val or ''}'. Dur: {h:02d}:{m:02d}:{s:02d}. Rec: {horas_rec_val_str or 'N/A'}"
        )
        limpiar_campos()
        regenerar_excel_desde_bd()
    except sqlite3.Error as e:
        messagebox.showerror(
            "Error de Base de Datos", f"No se pudo registrar la salida: {e}", parent=ventana
        )
        logger.error(f"Error BD salida {matricula}: {e}", exc_info=True)
    finally:
        conn.close()


def registrar_recuperacion_standalone_accion(evento=None):
    matricula = entrada_matricula.get().strip()
    if not (matricula.isdigit() and len(matricula) == 7):
        messagebox.showerror("Error de Entrada", "La matrícula debe ser de 7 números.", parent=ventana)
        return

    horas_str = entrada_horas_rec.get().strip()
    if not horas_str:
        messagebox.showerror(
            "Campos Requeridos", "Las horas a recuperar son requeridas.", parent=ventana
        )
        return
    try:
        horas_float = float(horas_str.replace(",", "."))
        if not (0 < horas_float <= 8):
            messagebox.showerror(
                "Entrada Inválida", "Las horas a recuperar deben estar entre 0.1 y 8.", parent=ventana
            )
            return
        horas_val_db = f"{horas_float:.1f}"
    except ValueError:
        messagebox.showerror(
            "Entrada Inválida", "Formato de horas a recuperar inválido.", parent=ventana
        )
        return

    try:
        fecha_falta_dt = entrada_fecha_falta_rec.get_date()
        fecha_falta_str = fecha_falta_dt.strftime("%d/%m/%Y")
    except ValueError:
        messagebox.showerror(
            "Error de Fecha", "Fecha de falta para recuperación inválida.", parent=ventana
        )
        return

    nota_val = entrada_nota.get().strip()
    if not nota_val:
        nota_val = None

    conn = obtener_conexion_bd()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT nombre FROM asesores WHERE matricula = ? AND activo = 1", (matricula,)
    )
    asesor = cursor.fetchone()
    if not asesor:
        messagebox.showerror(
            "Error de Matrícula",
            f"Matrícula '{matricula}' no encontrada o asesor inactivo.",
            parent=ventana
        )
        conn.close()
        return

    fecha_hoy_str = datetime.today().strftime("%Y-%m-%d")
    cursor.execute(
        "SELECT id FROM registros_asistencia WHERE matricula = ? AND fecha_registro = ? AND hora_salida IS NULL ORDER BY id DESC LIMIT 1",
        (matricula, fecha_hoy_str),
    )
    registro_abierto = cursor.fetchone()
    if not registro_abierto:
        messagebox.showerror(
            "Sin Entrada Abierta",
            "No se encontró una entrada abierta hoy para este asesor activo.",
            parent=ventana
        )
        conn.close()
        return

    try:
        update_fields = ["horas_recuperadas = ?", "fecha_falta_recuperada = ?"]
        params = [horas_val_db, fecha_falta_str]
        if nota_val is not None:
            update_fields.append("nota = ?")
            params.append(nota_val)

        query_str = (
            f"UPDATE registros_asistencia SET {', '.join(update_fields)} WHERE id = ?"
        )
        params.append(registro_abierto["id"])

        cursor.execute(query_str, tuple(params))
        conn.commit()
        crear_backup_bd_diario() # <--- AÑADIDO backup después de registrar recuperación
        mensaje_exito = f"Recuperación de {horas_val_db} hr(s) para {fecha_falta_str} registrada para {asesor['nombre']} (asociada a la entrada actual)."
        if nota_val:
            mensaje_exito += f"\nNota: {nota_val}"
        messagebox.showinfo("Recuperación Registrada", mensaje_exito, parent=ventana)
        logger.info(
            f"Rec (standalone): {horas_val_db}h for {fecha_falta_str} by {matricula}. Nota: '{nota_val or ''}'. ID reg {registro_abierto['id']}."
        )
        limpiar_campos()
        regenerar_excel_desde_bd()
    except sqlite3.Error as e:
        messagebox.showerror(
            "Error de Base de Datos", f"No se pudo registrar la recuperación: {e}", parent=ventana
        )
        logger.error(f"Error BD rec {matricula}: {e}", exc_info=True)
    finally:
        conn.close()


def importar_asesores_desde_excel_dialogo():
    if not messagebox.askyesno(
        "Confirmar Importación de Asesores",
        "Esto actualizará la lista de asesores activos.\n"
        "Los asesores no presentes en el archivo Excel serán marcados como INACTIVOS.\n"
        "Los asesores presentes en el archivo serán creados (si no existen) o actualizados y marcados como ACTIVOS.\n"
        "Los registros de asistencia existentes se conservarán.\n"
        "¿Desea continuar?",
        parent=ventana
    ):
        return

    ruta_archivo_maestro = filedialog.askopenfilename(
        title="Seleccionar Archivo Maestro de Asesores (Excel)",
        filetypes=(("Archivos Excel", "*.xlsx *.xls"), ("Todos los archivos", "*.*")),
        parent=ventana
    )
    if not ruta_archivo_maestro:
        return

    conn = None
    try:
        wb_maestro = openpyxl.load_workbook(ruta_archivo_maestro, read_only=True)
        if "Asesores" not in wb_maestro.sheetnames:
            messagebox.showerror(
                "Error de Hoja",
                "El archivo maestro de Excel debe contener una hoja llamada 'Asesores'.",
                parent=ventana
            )
            return
        ws_maestro = wb_maestro["Asesores"]

        conn = obtener_conexion_bd()
        cursor = conn.cursor()

        cursor.execute("UPDATE asesores SET activo = 0")
        logger.info("Todos los asesores existentes han sido marcados como inactivos.")

        insertados = 0
        actualizados_reactivados = 0
        cabeceras_excel = [celda.value for celda in ws_maestro[1]]
        cabeceras_esperadas_map = {
            "Nombre": None,
            "Matrícula": None,
            "Carrera": None,
            "Programa": None,
        }

        for cab_key in cabeceras_esperadas_map.keys():
            try:
                if cab_key == "Matrícula":
                    try:
                        cabeceras_esperadas_map[cab_key] = cabeceras_excel.index(
                            "Matrícula"
                        )
                    except ValueError:
                        cabeceras_esperadas_map[cab_key] = cabeceras_excel.index(
                            "Matricula"
                        )
                else:
                    cabeceras_esperadas_map[cab_key] = cabeceras_excel.index(cab_key)
            except ValueError:
                messagebox.showerror(
                    "Error de Formato de Cabecera",
                    f"La columna requerida '{cab_key}' no se encontró en la hoja 'Asesores'.",
                    parent=ventana
                )
                if conn:
                    conn.rollback()
                    conn.close()
                    return

        for num_fila, fila_valores in enumerate(
            ws_maestro.iter_rows(min_row=2, values_only=True), start=2
        ):
            try:
                matricula_val = fila_valores[cabeceras_esperadas_map["Matrícula"]]
                nombre_val = fila_valores[cabeceras_esperadas_map["Nombre"]]
                carrera_val = fila_valores[cabeceras_esperadas_map["Carrera"]]
                programa_val = fila_valores[cabeceras_esperadas_map["Programa"]]
            except IndexError:
                logger.warning(
                    f"Importación: Saltando fila {num_fila}, número incorrecto de columnas."
                )
                continue

            if not matricula_val or not (
                str(matricula_val).strip().isdigit()
                and len(str(matricula_val).strip()) == 7
            ):
                logger.warning(
                    f"Importación: Saltando fila {num_fila}, matrícula '{matricula_val}' inválida."
                )
                continue
            if not all([nombre_val, carrera_val, programa_val]):
                logger.warning(
                    f"Importación: Saltando fila {num_fila} para matrícula {matricula_val}, campos requeridos vacíos."
                )
                continue

            matricula_str = str(matricula_val).strip()
            nombre_str = str(nombre_val).strip()
            carrera_str = str(carrera_val).strip()
            programa_str = str(programa_val).strip()

            cursor.execute(
                "SELECT matricula FROM asesores WHERE matricula = ?", (matricula_str,)
            )
            asesor_existente = cursor.fetchone()
            if asesor_existente:
                cursor.execute(
                    "UPDATE asesores SET nombre = ?, carrera = ?, programa = ?, activo = 1 WHERE matricula = ?",
                    (nombre_str, carrera_str, programa_str, matricula_str),
                )
                actualizados_reactivados += 1
            else:
                cursor.execute(
                    "INSERT INTO asesores (matricula, nombre, carrera, programa, activo) VALUES (?, ?, ?, ?, 1)",
                    (matricula_str, nombre_str, carrera_str, programa_str),
                )
                insertados += 1

        conn.commit()
        crear_backup_bd_diario() # <--- AÑADIDO backup después de importar asesores
        messagebox.showinfo(
            "Importación Completa",
            f"{insertados} asesores nuevos importados y activados.\n"
            f"{actualizados_reactivados} asesores existentes actualizados y/o reactivados.",
            parent=ventana
        )
        logger.info(
            f"Importación: {insertados} nuevos activos, {actualizados_reactivados} actualizados/reactivados desde {ruta_archivo_maestro}"
        )
        regenerar_excel_desde_bd()
    except sqlite3.Error as e_sql:
        if conn:
            conn.rollback()
        messagebox.showerror(
            "Error de Base de Datos Durante Importación", f"Ocurrió un error: {e_sql}", parent=ventana
        )
        logger.error(f"Error SQLite importando asesores: {e_sql}", exc_info=True)
    except Exception as e:
        if conn:
            conn.rollback()
        messagebox.showerror(
            "Error de Importación", f"Ocurrió un error al importar asesores: {e}", parent=ventana
        )
        logger.error(f"Error importando asesores: {e}", exc_info=True)
    finally:
        if conn:
            conn.close()


def calcular_horas_mensuales_accion():
    matricula = entrada_matricula.get().strip()
    if not (matricula.isdigit() and len(matricula) == 7):
        messagebox.showerror(
            "Matrícula Requerida",
            "Por favor, ingrese la matrícula del asesor (7 números).",
            parent=ventana
        )
        return

    try:
        mes_str = entrada_mes_consulta.get()
        anio_str = entrada_anio_consulta.get()
        if not (mes_str.isdigit() and 1 <= int(mes_str) <= 12 and len(mes_str) <= 2):
            messagebox.showerror(
                "Entrada Inválida", "Mes debe ser un número entre 1 y 12 (ej: 7 o 07).", parent=ventana
            )
            return
        mes = int(mes_str)

        if not (anio_str.isdigit() and len(anio_str) == 4 and int(anio_str) >= 2024):
            messagebox.showerror(
                "Entrada Inválida",
                "Año debe ser un número de 4 dígitos desde 2024 en adelante.",
                parent=ventana
            )
            return
        anio = int(anio_str)
    except ValueError:
        messagebox.showerror("Entrada Inválida", "Mes o año con formato incorrecto.", parent=ventana)
        return

    conn = obtener_conexion_bd()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT nombre FROM asesores WHERE matricula = ? AND activo = 1", (matricula,)
    )
    asesor = cursor.fetchone()
    if not asesor:
        messagebox.showerror(
            "Asesor no Encontrado",
            f"No se encontró un asesor activo con matrícula '{matricula}'.",
            parent=ventana
        )
        conn.close()
        return

    nombre_asesor = asesor["nombre"]
    mes_fmt = f"{mes:02d}"
    primer_dia_mes = f"{anio}-{mes_fmt}-01"
    ultimo_dia_mes_num = calendar.monthrange(anio, mes)[1]
    ultimo_dia_mes = f"{anio}-{mes_fmt}-{ultimo_dia_mes_num:02d}"

    cursor.execute(
        "SELECT hora_entrada, hora_salida, horas_recuperadas FROM registros_asistencia WHERE matricula = ? AND fecha_registro BETWEEN ? AND ?",
        (matricula, primer_dia_mes, ultimo_dia_mes),
    )
    total_segundos_trabajados = 0
    total_horas_recuperadas = 0.0
    for registro in cursor.fetchall():
        if registro["hora_entrada"] and registro["hora_salida"]:
            try:
                dt_entrada = datetime.strptime(registro["hora_entrada"], "%H:%M:%S")
                dt_salida = datetime.strptime(registro["hora_salida"], "%H:%M:%S")
                diff = (dt_salida - dt_entrada).total_seconds()
                if diff < 0:
                    diff += 86400
                total_segundos_trabajados += diff
            except ValueError:
                logger.warning(
                    f"Formato hora inválido para {matricula} en {mes_fmt}/{anio}"
                )
        if registro["horas_recuperadas"]:
            try:
                valor_recuperadas_str = str(registro["horas_recuperadas"]).replace(
                    ",", "."
                )
                total_horas_recuperadas += float(valor_recuperadas_str)
            except ValueError:
                logger.warning(
                    f"Valor horas_recuperadas inválido para {matricula}: {registro['horas_recuperadas']}"
                )
    conn.close()

    h_trab = int(total_segundos_trabajados // 3600)
    m_trab = int((total_segundos_trabajados % 3600) // 60)
    s_trab = int(total_segundos_trabajados % 60)
    horas_trab_str = f"{h_trab:02d}:{m_trab:02d}:{s_trab:02d}"
    horas_rec_str = f"{total_horas_recuperadas:.1f}".replace(".", ",")

    resultado_msg = (
        f"Resumen para {nombre_asesor} (Matrícula: {matricula})\n"
        f"Mes: {mes_fmt}/{anio}\n\n"
        f"Horas Trabajadas (Entrada/Salida): {horas_trab_str}\n"
        f"Horas Recuperadas Registradas: {horas_rec_str} horas"
    )
    messagebox.showinfo("Horas Mensuales Calculadas", resultado_msg, parent=ventana)
    logger.info(
        f"Consulta horas: {matricula}, Mes: {mes_fmt}/{anio}. Trab: {horas_trab_str}, Rec: {horas_rec_str}h"
    )


# --- Función para Diálogo y Generación de Reporte Mensual Avanzado ---
def dialogo_generar_reporte_mensual_avanzado():
    ahora = datetime.now()
    mes_actual = ahora.month
    anio_actual = ahora.year

    mes = simpledialog.askinteger(
        "Reporte Mensual Avanzado",
        "Ingrese el número del mes (1-12):",
        parent=ventana,
        minvalue=1,
        maxvalue=12,
        initialvalue=mes_actual
    )
    if mes is None:
        return

    anio = simpledialog.askinteger(
        "Reporte Mensual Avanzado",
        f"Ingrese el año (ej: {anio_actual}):",
        parent=ventana,
        minvalue=2024,
        maxvalue=anio_actual + 5,
        initialvalue=anio_actual
    )
    if anio is None:
        return

    nombre_archivo_sugerido = f"ReporteMensualAvanzado_Asesores_{mes:02d}-{anio}.xlsx"
    ruta_archivo_reporte = filedialog.asksaveasfilename(
        title="Guardar Reporte Mensual Avanzado Como...",
        defaultextension=".xlsx",
        initialfile=nombre_archivo_sugerido,
        filetypes=(("Archivos Excel", ".xlsx"), ("Todos los archivos", ".*")),
        parent=ventana
    )
    if not ruta_archivo_reporte:
        return
    generar_reporte_mensual_avanzado(mes, anio, ruta_archivo_reporte)



def generar_reporte_mensual_avanzado(mes, anio, ruta_archivo_reporte):
    logger.info(
        f"Iniciando reporte mensual avanzado para {mes:02d}-{anio} en '{ruta_archivo_reporte}'"
    )
    wb_reporte = openpyxl.Workbook()
    wb_reporte.remove(wb_reporte.active)
    font_cabecera = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    fill_cabecera_resumen = PatternFill(
        start_color="2F75B5", end_color="2F75B5", fill_type="solid"
    )
    fill_cabecera_detalle = PatternFill(
        start_color="548235", end_color="548235", fill_type="solid"
    )
    alignment_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_izq_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    conn = obtener_conexion_bd()
    cursor = conn.cursor()

    hoja_resumen = wb_reporte.create_sheet(title=f"Resumen_{mes:02d}-{anio}")
    cabeceras_resumen = [
        "Programa",
        "Carrera",
        "Nombre Asesor",
        "Matrícula",
        "Total Horas Trabajadas",
        "Total Horas Recuperadas",
        "Total Días Trabajados",
        "Días con <= 3h o Faltas",
        "Días Cortos/Faltas Recuperados",
        "Días Cortos/Faltas No Recuperados",
    ]
    hoja_resumen.append(cabeceras_resumen)
    for col_idx, titulo in enumerate(cabeceras_resumen, 1):
        cell = hoja_resumen.cell(row=1, column=col_idx)
        cell.font = font_cabecera
        cell.fill = fill_cabecera_resumen
        cell.alignment = alignment_centro
        cell.border = thin_border

    hoja_detalle_cortos = wb_reporte.create_sheet(
        title=f"DetalleDiasCortos_{mes:02d}-{anio}"
    )
    cabeceras_detalle_cortos = [
        "Programa",
        "Carrera",
        "Nombre Asesor",
        "Fecha Día Corto/Falta",
        "Matrícula",
        "Entrada Original",
        "Salida Original",
        "Horas Trabajadas Día",
        "Nota Registro Original",
        "¿Recuperado?",
        "Horas Recuperadas (Total)",
        "Fecha(s) Recuperación",
        "Nota(s) Recuperación",
    ]
    hoja_detalle_cortos.append(cabeceras_detalle_cortos)
    for col_idx, titulo in enumerate(cabeceras_detalle_cortos, 1):
        cell = hoja_detalle_cortos.cell(row=1, column=col_idx)
        cell.font = font_cabecera
        cell.fill = fill_cabecera_detalle
        cell.alignment = alignment_centro
        cell.border = thin_border

    mes_fmt_sql = f"{mes:02d}"
    primer_dia_mes_sql = f"{anio}-{mes_fmt_sql}-01"
    ultimo_dia_mes_num = calendar.monthrange(anio, mes)[1]
    ultimo_dia_mes_sql = f"{anio}-{mes_fmt_sql}-{ultimo_dia_mes_num:02d}"

    cursor.execute(
        "SELECT DISTINCT fecha_registro FROM registros_asistencia WHERE fecha_registro BETWEEN ? AND ? ORDER BY fecha_registro",
        (primer_dia_mes_sql, ultimo_dia_mes_sql),
    )
    dias_laborables_inferidos_obj = {
        datetime.strptime(row["fecha_registro"], "%Y-%m-%d")
        for row in cursor.fetchall()
    }

    cursor.execute(
        "SELECT matricula, nombre, programa, carrera FROM asesores WHERE activo = 1 ORDER BY programa, carrera, nombre"
    )
    lista_asesores = cursor.fetchall()

    fila_actual_resumen = 2
    fila_actual_detalle_cortos = 2

    for asesor_actual in lista_asesores:
        matricula_asesor = asesor_actual["matricula"]
        nombre_asesor = asesor_actual["nombre"]
        programa_asesor = asesor_actual["programa"]
        carrera_asesor = asesor_actual["carrera"]

        total_segundos_trabajados_mes = 0
        total_horas_recuperadas_mes = 0.0
        dias_trabajados_por_asesor_obj = set()
        dias_cortos_info = []

        cursor.execute(
            "SELECT fecha_registro, hora_entrada, hora_salida, horas_recuperadas, nota FROM registros_asistencia WHERE matricula = ? AND fecha_registro BETWEEN ? AND ?",
            (matricula_asesor, primer_dia_mes_sql, ultimo_dia_mes_sql),
        )
        registros_mes_asesor = cursor.fetchall()
        registros_asesor_por_fecha_str = {
            reg["fecha_registro"]: reg for reg in registros_mes_asesor
        }

        for fecha_laborable_obj in sorted(list(dias_laborables_inferidos_obj)):
            fecha_laborable_str = fecha_laborable_obj.strftime("%Y-%m-%d")
            if fecha_laborable_str in registros_asesor_por_fecha_str:
                reg = registros_asesor_por_fecha_str[fecha_laborable_str]
                dias_trabajados_por_asesor_obj.add(fecha_laborable_obj)
                segundos_dia = 0
                if reg["hora_entrada"] and reg["hora_salida"]:
                    try:
                        dt_e = datetime.strptime(reg["hora_entrada"], "%H:%M:%S")
                        dt_s = datetime.strptime(reg["hora_salida"], "%H:%M:%S")
                        diff = (dt_s - dt_e).total_seconds()
                        if diff < 0:
                            diff += 86400
                        total_segundos_trabajados_mes += diff
                        segundos_dia = diff
                    except ValueError:
                        logger.warning(
                            f"Error parseando horas para {matricula_asesor} el {reg['fecha_registro']}"
                        )
                if reg["horas_recuperadas"]:
                    try:
                        total_horas_recuperadas_mes += float(
                            str(reg["horas_recuperadas"]).replace(",", ".")
                        )
                    except ValueError:
                        pass
                if segundos_dia <= (3 * 3600):
                    dias_cortos_info.append(
                        (fecha_laborable_obj, segundos_dia, reg["nota"], "Pocas Horas")
                    )
            else:
                dias_cortos_info.append(
                    (fecha_laborable_obj, 0, "Falta (Sin Registro)", "Falta")
                )

        h_trab_mes = int(total_segundos_trabajados_mes // 3600)
        m_trab_mes = int((total_segundos_trabajados_mes % 3600) // 60)
        s_trab_mes = int(total_segundos_trabajados_mes % 60)
        total_horas_trab_mes_str = f"{h_trab_mes:02d}:{m_trab_mes:02d}:{s_trab_mes:02d}"
        total_horas_rec_mes_str = f"{total_horas_recuperadas_mes:.1f}".replace(".", ",")
        total_dias_trabajados_mes = len(dias_trabajados_por_asesor_obj)
        num_dias_cortos_o_faltas = len(dias_cortos_info)
        num_dias_cortos_o_faltas_recuperados = 0

        for (
            dia_data
        ) in (
            dias_cortos_info
        ):
            fecha_dia_obj, segundos_trab_dia, nota_original_dia, tipo_dia = dia_data
            fecha_dia_str_excel = fecha_dia_obj.strftime("%d-%m-%Y")
            fecha_dia_str_sql_recup = fecha_dia_obj.strftime("%d/%m/%Y")
            h_dc = int(segundos_trab_dia // 3600)
            m_dc = int((segundos_trab_dia % 3600) // 60)
            s_dc = int(segundos_trab_dia % 60)
            horas_trab_dia_str = f"{h_dc:02d}:{m_dc:02d}:{s_dc:02d}"
            entrada_orig_dc = ""
            salida_orig_dc = ""
            if tipo_dia != "Falta":
                fecha_dia_str_original_sql = fecha_dia_obj.strftime("%Y-%m-%d")
                if fecha_dia_str_original_sql in registros_asesor_por_fecha_str:
                    reg_original = registros_asesor_por_fecha_str[
                        fecha_dia_str_original_sql
                    ]
                    entrada_orig_dc = (
                        reg_original["hora_entrada"] if reg_original["hora_entrada"] else ""
                    )
                    salida_orig_dc = reg_original["hora_salida"] if reg_original["hora_salida"] else ""

            cursor.execute(
                "SELECT SUM(CAST(REPLACE(horas_recuperadas, ',', '.') AS REAL)) as total_rec, GROUP_CONCAT(fecha_registro, '; ') as fechas_rec, GROUP_CONCAT(IFNULL(nota, ''), ' | ') as notas_rec FROM registros_asistencia WHERE matricula = ? AND fecha_falta_recuperada = ? AND horas_recuperadas IS NOT NULL",
                (matricula_asesor, fecha_dia_str_sql_recup),
            )
            info_recuperacion = cursor.fetchone()
            se_recupero_str = "No"
            horas_rec_para_falta = 0.0
            fechas_de_recuperacion = ""
            notas_de_recuperacion = ""
            if (
                info_recuperacion
                and info_recuperacion["total_rec"] is not None
                and info_recuperacion["total_rec"] > 0
            ):
                se_recupero_str = "Sí"
                num_dias_cortos_o_faltas_recuperados += 1
                horas_rec_para_falta = info_recuperacion["total_rec"]
                fechas_de_recuperacion = (
                    info_recuperacion["fechas_rec"]
                    if info_recuperacion["fechas_rec"]
                    else ""
                )
                notas_de_recuperacion = (
                    info_recuperacion["notas_rec"].strip(" | ")
                    if info_recuperacion["notas_rec"]
                    else ""
                )
            
            datos_fila_detalle = [
                programa_asesor,
                carrera_asesor,
                nombre_asesor,
                fecha_dia_str_excel,
                matricula_asesor,
                entrada_orig_dc,
                salida_orig_dc,
                horas_trab_dia_str,
                nota_original_dia if nota_original_dia else "",
                se_recupero_str,
                f"{horas_rec_para_falta:.1f}".replace(".", ","),
                fechas_de_recuperacion,
                notas_de_recuperacion,
            ]
            hoja_detalle_cortos.append(datos_fila_detalle)
            for col_idx_detalle, valor_celda in enumerate(datos_fila_detalle, 1):
                cell = hoja_detalle_cortos.cell(
                    row=fila_actual_detalle_cortos, column=col_idx_detalle
                )
                cell.border = thin_border
                cell.alignment = (
                    alignment_izq_wrap
                    if col_idx_detalle in [9, 12, 13] 
                    else alignment_centro
                )
            fila_actual_detalle_cortos += 1

        num_dias_cortos_o_faltas_no_recuperados = (
            num_dias_cortos_o_faltas - num_dias_cortos_o_faltas_recuperados
        )
        datos_fila_resumen = [
            programa_asesor,
            carrera_asesor,
            nombre_asesor,
            matricula_asesor,
            total_horas_trab_mes_str,
            total_horas_rec_mes_str,
            total_dias_trabajados_mes,
            num_dias_cortos_o_faltas,
            num_dias_cortos_o_faltas_recuperados,
            num_dias_cortos_o_faltas_no_recuperados,
        ]
        hoja_resumen.append(datos_fila_resumen)
        for col_idx_resumen, valor_celda in enumerate(datos_fila_resumen, 1):
            cell = hoja_resumen.cell(row=fila_actual_resumen, column=col_idx_resumen)
            cell.border = thin_border
            cell.alignment = (
                alignment_izq_wrap if col_idx_resumen == 3 else alignment_centro
            )
        fila_actual_resumen += 1
    conn.close()

    for hoja in [hoja_resumen, hoja_detalle_cortos]:
        for col in hoja.columns:
            max_l = 0
            col_letra = col[0].column_letter
            for celda in col:
                try:
                    if celda.value:
                        val_str = str(celda.value)
                        if (
                            celda.alignment
                            and celda.alignment.wrap_text
                            and ("\n" in val_str or "; " in val_str or " | " in val_str)
                        ):
                            l_celda = max(
                                len(s)
                                for s in val_str.replace("; ", "\n")
                                .replace(" | ", "\n")
                                .split("\n")
                            )
                        else:
                            l_celda = len(val_str)
                        max_l = max(max_l, l_celda)
                except:
                    pass
            ancho = (max_l + 4) if max_l > 0 else 12
            
            if hoja.title.startswith("Detalle"):
                if col_letra in ["I", "M"]:
                    ancho = max(ancho, 35)
                elif col_letra == "L":
                    ancho = max(ancho, 30)
            hoja.column_dimensions[col_letra].width = ancho
    try:
        wb_reporte.save(ruta_archivo_reporte)
        messagebox.showinfo(
            "Reporte Generado",
            f"El reporte mensual avanzado ha sido guardado en:\n{ruta_archivo_reporte}",
            parent=ventana
        )
        logger.info(
            f"Reporte mensual avanzado generado y guardado: {ruta_archivo_reporte}"
        )
    except PermissionError:
        messagebox.showerror(
            "Error al Guardar",
            f"Permiso denegado al guardar reporte en '{ruta_archivo_reporte}'.\nAsegúrese de que el archivo no esté abierto o la ubicación sea escribible.",
            parent=ventana
        )
        logger.error(
            f"PermissionError al guardar reporte mensual: {ruta_archivo_reporte}"
        )
    except Exception as e:
        messagebox.showerror(
            "Error Inesperado", f"No se pudo generar o guardar el reporte mensual: {e}", parent=ventana
        )
        logger.error(f"Error generando reporte mensual: {e}", exc_info=True)

        
# --- GUI Setup ---
ventana = tk.Tk()
ventana.title("Sistema de Registro de Asistencia de Asesores")
ventana.geometry("650x450")
ventana.configure(bg="#F0F0F0")

vcmd_matricula = (ventana.register(lambda P: validar_solo_numeros_longitud(P, 7)), "%P")
vcmd_horas_rec = (ventana.register(validar_horas_recuperar), "%P")
vcmd_mes = (ventana.register(validar_mes), "%P")
vcmd_anio = (ventana.register(validar_anio), "%P")

barra_menu = tk.Menu(ventana)
menu_administracion = tk.Menu(barra_menu, tearoff=0)
menu_administracion.add_command(
    label="Importar/Sobrescribir Lista de Asesores...",
    command=importar_asesores_desde_excel_dialogo,
)
menu_administracion.add_command(
    label="Generar Reporte Mensual Avanzado...",
    command=dialogo_generar_reporte_mensual_avanzado,
)
menu_administracion.add_separator() # <--- AÑADIDO Separador
menu_administracion.add_command(
    label="Restaurar Base de Datos desde Backup...", # <--- NUEVA OPCIÓN
    command=dialogo_restaurar_bd_desde_backup
)
barra_menu.add_cascade(label="Administración", menu=menu_administracion)
ventana.config(menu=barra_menu)

fuente_etiqueta = ("Segoe UI", 10)
fuente_entrada = ("Segoe UI", 10)
fuente_boton = ("Segoe UI", 10, "bold")
color_fondo_frame = "#F0F0F0"
color_etiqueta_fondo = color_fondo_frame
color_boton_entrada_fondo = "#4CAF50"
color_boton_entrada_texto = "white"
color_boton_salida_fondo = "#FF9800"
color_boton_salida_texto = "white"
color_boton_recup_fondo = "#2196F3"
color_boton_recup_texto = "white"
color_boton_accion_fondo = "#0078D4"
color_boton_accion_texto = "white"

frame_principal = tk.Frame(ventana, bg=color_fondo_frame, padx=10, pady=5)
frame_principal.pack(fill="x")
tk.Label(
    frame_principal,
    text="Matrícula del Asesor (7 números):",
    font=fuente_etiqueta,
    bg=color_etiqueta_fondo,
).grid(row=0, column=0, sticky="w", pady=(0, 5))
entrada_matricula = tk.Entry(
    frame_principal,
    font=fuente_entrada,
    width=20,
    validate="key",
    validatecommand=vcmd_matricula,
)
entrada_matricula.grid(row=0, column=1, sticky="ew", pady=(0, 5))
frame_principal.grid_columnconfigure(1, weight=1)

frame_botones_principales = tk.Frame(ventana, bg=color_fondo_frame, padx=10)
frame_botones_principales.pack(fill="x")
boton_entrada = tk.Button(
    frame_botones_principales,
    text="Registrar Entrada (Enter)",
    font=fuente_boton,
    bg=color_boton_entrada_fondo,
    fg=color_boton_entrada_texto,
    command=registrar_entrada_accion,
)
boton_entrada.pack(side="left", fill="x", expand=True, padx=(0, 5), pady=5)
boton_salida = tk.Button(
    frame_botones_principales,
    text="Registrar Salida (Shift)",
    font=fuente_boton,
    bg=color_boton_salida_fondo,
    fg=color_boton_salida_texto,
    command=lambda: registrar_salida_accion(None),
)
boton_salida.pack(side="left", fill="x", expand=True, padx=(5, 0), pady=5)

frame_nota = tk.Frame(ventana, bg=color_fondo_frame, padx=10)
frame_nota.pack(fill="x", pady=(5, 0))
tk.Label(
    frame_nota, text="Nota (Opcional):", font=fuente_etiqueta, bg=color_etiqueta_fondo
).pack(side="left", padx=(0, 5))
entrada_nota = tk.Entry(frame_nota, font=fuente_entrada)
entrada_nota.pack(side="left", fill="x", expand=True)

frame_recuperacion = tk.LabelFrame(
    ventana,
    text=" Horas de Recuperación (Opcional, máx. 8h) ",
    font=("Segoe UI", 9, "bold"),
    bg=color_fondo_frame,
    padx=10,
    pady=5,
    relief=tk.GROOVE,
    borderwidth=1,
)
frame_recuperacion.pack(fill="x", padx=10)
tk.Label(
    frame_recuperacion,
    text="Horas a Recuperar (ej: 1, 1.5):",
    font=fuente_etiqueta,
    bg=color_etiqueta_fondo,
).grid(row=0, column=0, sticky="w", pady=2)
entrada_horas_rec = tk.Entry(
    frame_recuperacion,
    font=fuente_entrada,
    width=8,
    validate="key",
    validatecommand=vcmd_horas_rec,
)
entrada_horas_rec.grid(row=0, column=1, sticky="w", padx=5, pady=2)
tk.Label(
    frame_recuperacion,
    text="Fecha de Falta (para la cual se recupera):",
    font=fuente_etiqueta,
    bg=color_etiqueta_fondo,
).grid(row=1, column=0, sticky="w", pady=2)
entrada_fecha_falta_rec = DateEntry(
    frame_recuperacion,
    font=fuente_entrada,
    width=12,
    date_pattern="dd/mm/yyyy",
    state="readonly",
    maxdate=datetime.today() - timedelta(days=1),
    locale="es_MX",
)
entrada_fecha_falta_rec.grid(row=1, column=1, sticky="w", padx=5, pady=2)
boton_recuperacion_standalone = tk.Button(
    frame_recuperacion,
    text="Registrar Solo Recuperación\n(Asociar a Entrada Actual)",
    font=fuente_boton,
    bg=color_boton_recup_fondo,
    fg=color_boton_recup_texto,
    command=registrar_recuperacion_standalone_accion,
    justify=tk.CENTER,
)
boton_recuperacion_standalone.grid(
    row=0, column=2, rowspan=2, sticky="nsew", padx=(10, 0), pady=2
)
frame_recuperacion.grid_columnconfigure(2, weight=1)

frame_consulta = tk.LabelFrame(
    ventana,
    text=" Consulta de Horas Mensuales por Asesor ",
    font=("Segoe UI", 9, "bold"),
    bg=color_fondo_frame,
    padx=10,
    pady=5,
    relief=tk.GROOVE,
    borderwidth=1,
)
frame_consulta.pack(fill="x", padx=10)
tk.Label(
    frame_consulta, text="Mes (1-12):", font=fuente_etiqueta, bg=color_etiqueta_fondo
).grid(row=0, column=0, sticky="w", pady=2)
entrada_mes_consulta = tk.Entry(
    frame_consulta,
    font=fuente_entrada,
    width=5,
    validate="key",
    validatecommand=vcmd_mes,
)
entrada_mes_consulta.insert(0, str(datetime.now().month))
entrada_mes_consulta.grid(row=0, column=1, sticky="w", padx=5, pady=2)
tk.Label(
    frame_consulta, text="Año (YYYY):", font=fuente_etiqueta, bg=color_etiqueta_fondo
).grid(row=0, column=2, sticky="w", padx=(10, 0), pady=2)
entrada_anio_consulta = tk.Entry(
    frame_consulta,
    font=fuente_entrada,
    width=7,
    validate="key",
    validatecommand=vcmd_anio,
)
entrada_anio_consulta.insert(0, str(datetime.now().year))
entrada_anio_consulta.grid(row=0, column=3, sticky="w", padx=5, pady=2)
boton_calcular_horas = tk.Button(
    frame_consulta,
    text="Calcular Horas",
    font=fuente_boton,
    bg=color_boton_accion_fondo,
    fg=color_boton_accion_texto,
    command=calcular_horas_mensuales_accion,
)
boton_calcular_horas.grid(row=0, column=4, sticky="ew", padx=(10, 0), pady=2)
frame_consulta.grid_columnconfigure(4, weight=1)

frame_actualizar_excel = tk.Frame(ventana, bg=color_fondo_frame, padx=10, pady=5)
frame_actualizar_excel.pack(fill="x")
boton_actualizar_excel = tk.Button(
    frame_actualizar_excel,
    text="Actualizar Reporte de Asistencias Manualmente",
    font=fuente_boton,
    bg=color_boton_accion_fondo,
    fg=color_boton_accion_texto,
    command=lambda: regenerar_excel_desde_bd(mostrar_mensaje_exito=True),
)
boton_actualizar_excel.pack(fill="x")

ventana.bind("<Return>", registrar_entrada_accion)
ventana.bind("<KP_Enter>", registrar_entrada_accion)
ventana.bind("<Shift_L>", registrar_salida_accion)
ventana.bind("<Shift_R>", registrar_salida_accion)

if __name__ == "__main__":
    inicializar_bd() # Esto ya llama a crear_backup_bd_diario()
    if not os.path.exists(NOMBRE_ARCHIVO_EXCEL):
        logger.info(
            f"Archivo Excel '{NOMBRE_ARCHIVO_EXCEL}' no encontrado. Generando al inicio."
        )
        regenerar_excel_desde_bd(mostrar_mensaje_exito=False)
    
    # Asegurar que el directorio de backups exista al inicio
    if not os.path.exists(DIRECTORIO_BACKUPS_BD):
        try:
            os.makedirs(DIRECTORIO_BACKUPS_BD)
            logger.info(f"Directorio de backups '{DIRECTORIO_BACKUPS_BD}' creado al inicio.")
        except OSError as e:
            logger.error(f"No se pudo crear el directorio de backups '{DIRECTORIO_BACKUPS_BD}' al inicio: {e}")
            # No es fatal, la función de backup lo intentará crear de nuevo

    entrada_matricula.focus_set()
    ventana.mainloop()