import tkinter as tk
from tkinter import messagebox
from datetime import datetime, timedelta
import openpyxl
import os
import logging
import sys
from tkcalendar import DateEntry

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    filename='app.log'
)

logger = logging.getLogger(__name__)

def capturar_excepcion(exc_type, exc_value, exc_traceback):
    logger.error("Excepcion no capturada", exc_info=(exc_type, exc_value, exc_traceback))


sys.excepthook = capturar_excepcion

def borrar_inputs():
    entrada_matricula.delete(0, tk.END)
    entrada_horas.delete(0, tk.END)
    entrada_fecha_falta.set_date(datetime.today() - timedelta(days=1))


def cargar_excel():
    """Carga el archivo Excel o lo crea si no existe."""
    archivo = "Asesores.xlsx"
    if not os.path.exists(archivo):
        messagebox.showerror(
            "Error",
            "El archivo Asesores.xlsx no existe. Por favor, crea el archivo antes de continuar. El archivo debe contener una hoja llamada 'Asesores' con los datos de los asesores (Nombre, Matricula, Carrera, Programa).",
        )
        return
    return openpyxl.load_workbook(archivo)


def inicializar_excel():
    """Verifica si existe la hoja del día actual y la crea si no está."""
    wb = cargar_excel()
    fecha_hoy = datetime.today().strftime("%Y-%m-%d")
    if fecha_hoy not in wb.sheetnames:
        ws_dia = wb.create_sheet(title=fecha_hoy)
        ws_dia.append(
            [
                "Nombre",
                "Matrícula",
                "Hora de Entrada",
                "Hora de Salida",
                "Horas Recuperadas",
                "Fecha de Falta",
            ]
        )
        try:
            wb.save("Asesores.xlsx")
        except PermissionError:
            messagebox.showerror(
                "Error",
                "Acceso denegado, asegurate de que el archivo Asesores.xlsx no este abierto",
            )
            return


def registrar_entrada(event):
    matricula = entrada_matricula.get().strip()
    inicializar_excel()  # Asegurar que la hoja del día existe
    wb = cargar_excel()
    ws_asesores = wb["Asesores"]

    asesor = None
    for row in ws_asesores.iter_rows(min_row=2, values_only=True):
        if str(row[1]).strip() == matricula:
            asesor = row
            break

    if not asesor:
        messagebox.showerror("Error", f"Matrícula '{matricula}' no encontrada")
        return

    fecha_hoy = datetime.today().strftime("%Y-%m-%d")
    ws_dia = wb[fecha_hoy]

    for row in ws_dia.iter_rows(min_row=2):
        if str(row[1].value).strip() == matricula and not row[3].value:
            messagebox.showwarning(
                "Aviso",
                "Existe una entrada para este asesor sin registro de salida. Favor de registrar la salida primero.",
            )
            return

    hora_entrada = datetime.now().strftime("%H:%M:%S")
    ws_dia.append([asesor[0], matricula, hora_entrada, "", "", ""])

    try:
        wb.save("Asesores.xlsx")
    except PermissionError:
        messagebox.showerror(
            "Error",
            "Acceso denegado, asegurate de que el archivo Asesores.xlsx no este abierto",
        )
        return

    messagebox.showinfo(
        "Éxito", f"Entrada registrada para {asesor[0]} ({asesor[2]}) a las {hora_entrada}"
    )

    if entrada_horas.get().strip():
        registrar_recuperacion(event)
        
    borrar_inputs()


def registrar_salida(event):
    matricula = entrada_matricula.get().strip()

    if entrada_horas.get().strip():
        registrar_recuperacion(event)

    wb = cargar_excel()
    inicializar_excel()  # Asegurar que la hoja del día existe
    fecha_hoy = datetime.today().strftime("%Y-%m-%d")
    ws_dia = wb[fecha_hoy]

    for row in ws_dia.iter_rows(min_row=2):
        if str(row[1].value).strip() == matricula and not row[3].value:
            row[3].value = datetime.now().strftime("%H:%M:%S")

            try:
                wb.save("Asesores.xlsx")
                diff = (datetime.strptime(row[3].value, '%H:%M:%S') - datetime.strptime(row[2].value, '%H:%M:%S')).total_seconds()
                horas = diff // 3600
                mins = (diff % 3600) // 60
                segs = diff % 60
                messagebox.showinfo("Éxito", f"Salida registrada con éxito para {row[0].value}, realizó {int(horas):02}:{int(mins):02}:{int(segs):02} horas")
                borrar_inputs()
                return
            except PermissionError:
                messagebox.showerror(
                    "Error",
                    "Acceso denegado, asegurate de que el archivo Asesores.xlsx no este abierto",
                )
                return

    messagebox.showerror(
        "Error", "No se encontró una entrada pendiente para esta matrícula"
    )


def registrar_recuperacion(event):
    matricula = entrada_matricula.get().strip()
    horas = entrada_horas.get().strip()
    fecha_falta = entrada_fecha_falta.get().strip()
    wb = cargar_excel()
    inicializar_excel()  # Asegurar que la hoja del día existe
    fecha_hoy = datetime.today().strftime("%Y-%m-%d")
    ws_dia = wb[fecha_hoy]

    for row in ws_dia.iter_rows(min_row=2):
        if str(row[1].value).strip() == matricula and not row[3].value:
            row[4].value = horas
            row[5].value = fecha_falta

            try:
                wb.save("Asesores.xlsx")
                messagebox.showinfo("Éxito", "Recuperación registrada correctamente.")
                borrar_inputs()
                return
            except PermissionError:
                messagebox.showerror(
                    "Error",
                    "Acceso denegado, asegurate de que el archivo Asesores.xlsx no este abierto",
                )
                return

    messagebox.showerror(
        "Error",
        "No se pudo encontrar un registro abierto del asesor para las horas de recuperacion. Favor de registrar entrada sin registrar salida primero",
    )


# Interfaz gráfica
root = tk.Tk()
root.title("Gestión de Asesores")
root.geometry("400x500")
root.configure(bg="#f0f0f0")

# Atajos de teclado
root.bind("<Return>", registrar_entrada)
root.bind("<KP_Enter>", registrar_entrada)
root.bind("<Shift_R>", registrar_salida)
root.bind("<Alt_L>", registrar_recuperacion)

# Entrada de matrícula
tk.Label(
    root, text="Ingrese Matrícula:", font=("Arial", 12, "bold"), bg="#f0f0f0"
).pack(pady=5)
entrada_matricula = tk.Entry(root, font=("Arial", 12))
entrada_matricula.config(
    validate="key",
    validatecommand=(
        root.register(lambda P: (P.isdigit() or P == "") and len(P) <= 7),
        "%P",
    ),
)
entrada_matricula.pack(pady=5)

# Boton de registrar entrada
btn_entrada = tk.Button(
    root, text="Registrar Entrada (Enter)", font=("Arial", 12), bg="#4CAF50", fg="white"
)
btn_entrada.bind("<Button-1>", registrar_entrada)
btn_entrada.pack(pady=5, fill=tk.X)

# Boton de registrar salida
btn_salida = tk.Button(
    root,
    text="Registrar Salida (RShift)",
    font=("Arial", 12),
    bg="#FF9800",
    fg="white",
)
btn_salida.bind("<Button-1>", registrar_salida)
btn_salida.pack(pady=5, fill=tk.X)

# Entrada de horas a recuperar
tk.Label(
    root, text="Ingrese Horas a Recuperar:", font=("Arial", 12, "bold"), bg="#f0f0f0"
).pack(pady=5)
entrada_horas = tk.Entry(root, font=("Arial", 12))
entrada_horas.config(
    validate="key",
    validatecommand=(
        root.register(lambda P: (P.isdigit() or P == "") and len(P) <= 1),
        "%P",
    ),
)
entrada_horas.pack(pady=5)

# Entrada de fecha de falta
tk.Label(
    root, text="Ingrese Fecha de Falta:", font=("Arial", 12, "bold"), bg="#f0f0f0"
).pack(pady=5)
entrada_fecha_falta = DateEntry(
    root,
    font=("Arial", 12),
    date_pattern="dd/mm/yyyy",
    state="readonly",
    maxdate=datetime.today() - timedelta(days=1),
    locale="es",
)
entrada_fecha_falta.pack(pady=5)

# Boton de registrar recuperacion
btn_recuperar = tk.Button(
    root,
    text="Registrar Recuperación (LAlt)",
    font=("Arial", 12),
    bg="#2196F3",
    fg="white",
)
btn_recuperar.bind("<Button-1>", registrar_recuperacion)
btn_recuperar.pack(pady=5, fill=tk.X)

# Iniciar la aplicación
root.mainloop()
